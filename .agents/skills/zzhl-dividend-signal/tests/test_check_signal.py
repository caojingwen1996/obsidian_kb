import sys
import tempfile
import unittest
import json
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from scripts import check_signal


class SelectLatestOnOrBeforeTest(unittest.TestCase):
    def test_selects_latest_row_on_or_before_target_date(self):
        frame = pd.DataFrame(
            {
                "日期": ["2026-07-01", "2026-07-02", "2026-07-03"],
                "股息率2": ["4.61", "4.64", "4.67"],
            }
        )

        row = check_signal.select_latest_on_or_before(
            frame,
            "日期",
            datetime(2026, 7, 2),
        )

        self.assertEqual(str(row["日期"])[:10], "2026-07-02")
        self.assertEqual(float(row["股息率2"]), 4.64)


class LixingerPublicPercentileParseTest(unittest.TestCase):
    def test_parses_public_page_percentile_fields(self):
        text = """
        <h1>中证红利(000922) - 历史股息率(市值加权)走势图</h1>
        <div>最后更新于：2026-07-03</div>
        <ul>
          <li>当前值:4.66%</li>
          <li>当前分位点 31.31%</li>
          <li>80%分位点 6.12%</li>
        </ul>
        """

        parsed = check_signal.parse_lixinger_public_percentile(
            text,
            "https://example.test/dyr?metrics-type=mcw",
        )

        self.assertEqual(parsed.date, "2026-07-03")
        self.assertEqual(parsed.dividend_yield, 4.66)
        self.assertEqual(parsed.percentile_10y, 31.31)
        self.assertEqual(parsed.percentile_80_value, 6.12)
        self.assertEqual(parsed.source, "理杏仁公开页面")


class LlmLixingerFallbackTest(unittest.TestCase):
    def test_uses_llm_query_values_when_public_page_fails(self):
        llm_percentile = check_signal.make_llm_lixinger_percentile(
            date="2026-07-03",
            dividend_yield="4.66",
            percentile_10y="30.44",
            percentile_80_value="6.12",
            source_note="Codex网页查询",
        )

        with patch.object(
            check_signal,
            "fetch_lixinger_public_percentile",
            side_effect=RuntimeError("HTTP Error 429: Too Many Requests"),
        ):
            parsed = check_signal.fetch_lixinger_percentile(llm_percentile)

        self.assertEqual(parsed.date, "2026-07-03")
        self.assertEqual(parsed.dividend_yield, 4.66)
        self.assertEqual(parsed.percentile_10y, 30.44)
        self.assertEqual(parsed.percentile_80_value, 6.12)
        self.assertEqual(parsed.source, "大模型网页查询")
        self.assertIn("公开页面失败后采用", parsed.note)


class XueqiuQuoteParseTest(unittest.TestCase):
    def test_parses_xueqiu_realtime_quote_payload(self):
        payload = """
        {
          "data": [
            {
              "symbol": "SH000922",
              "percent": -0.73,
              "timestamp": 1783319400000
            }
          ]
        }
        """

        parsed = check_signal.parse_xueqiu_quote_payload(payload)

        self.assertEqual(parsed.date, "2026-07-06")
        self.assertEqual(parsed.change_percent, -0.73)
        self.assertEqual(parsed.source, "雪球实时行情")


class AkshareDataSourceTest(unittest.TestCase):
    def test_uses_akshare_index_dividend_yield(self):
        class AkStub:
            def stock_zh_index_value_csindex(self, symbol):
                self.symbol = symbol
                return pd.DataFrame(
                    {
                        "日期": ["2026-07-02", "2026-07-03"],
                        "股息率2": ["4.64", "4.67"],
                    }
                )

        valuation = check_signal.fetch_index_valuation(
            AkStub(),
            datetime(2026, 7, 4),
        )

        self.assertEqual(valuation.date, "2026-07-03")
        self.assertEqual(valuation.dividend_yield, 4.67)
        self.assertIn("AKShare", valuation.source)
        self.assertNotIn("Tushare MCP", valuation.note)

    def test_calculates_calendar_year_return_and_max_drawdown(self):
        class AkStub:
            def stock_zh_index_daily(self, symbol):
                self.symbol = symbol
                return pd.DataFrame(
                    {
                        "date": [
                            "2025-12-31",
                            "2026-01-05",
                            "2026-01-06",
                            "2026-01-07",
                            "2026-01-08",
                        ],
                        "close": [98, 100, 110, 88, 105],
                    }
                )

        performance = check_signal.fetch_index_performance(
            AkStub(),
            datetime(2026, 1, 8),
        )

        self.assertEqual(performance.date, "2026-01-08")
        self.assertEqual(performance.start_date, "2026-01-05")
        self.assertAlmostEqual(performance.ytd_return, 5.0)
        self.assertAlmostEqual(performance.ytd_max_drawdown, -20.0)
        self.assertIn("stock_zh_index_daily", performance.note)

    def test_calculates_each_year_from_index_inception_to_current_year(self):
        frame = pd.DataFrame(
            {
                "date": [
                    "2008-05-26",
                    "2008-06-02",
                    "2008-12-31",
                    "2009-01-05",
                    "2009-06-01",
                    "2009-09-01",
                    "2009-12-31",
                    "2026-01-05",
                    "2026-01-06",
                    "2026-01-07",
                ],
                "close": [100, 80, 120, 100, 150, 90, 120, 100, 110, 99],
            }
        )

        annual = check_signal.calculate_annual_performance(
            frame,
            "date",
            "close",
            datetime(2026, 1, 7),
        )

        self.assertEqual([row.year for row in annual], [2008, 2009, 2026])
        self.assertEqual(annual[0].status, "成立首年")
        self.assertAlmostEqual(annual[0].annual_return, 20.0)
        self.assertAlmostEqual(annual[0].max_drawdown, -20.0)
        self.assertEqual(annual[1].status, "完整年度")
        self.assertAlmostEqual(annual[1].annual_return, 20.0)
        self.assertAlmostEqual(annual[1].max_drawdown, -40.0)
        self.assertEqual(annual[2].status, "年内")
        self.assertAlmostEqual(annual[2].annual_return, -1.0)
        self.assertAlmostEqual(annual[2].max_drawdown, -10.0)

    def test_writes_annual_performance_json(self):
        annual = [
            check_signal.AnnualPerformance(
                year=2026,
                start_date="2026-01-05",
                end_date="2026-08-19",
                annual_return=0.51,
                max_drawdown=-15.96,
                status="年内",
            )
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "annual.json"
            check_signal.write_annual_performance_json(
                annual,
                path,
                source="中证指数官网 index-perf(000922)",
            )
            payload = json.loads(path.read_text(encoding="utf-8"))

        self.assertEqual(payload["indexCode"], "000922")
        self.assertEqual(payload["firstDate"], "2026-01-05")
        self.assertEqual(payload["lastDate"], "2026-08-19")
        self.assertEqual(payload["rows"][0]["status"], "年内")

    def test_uses_akshare_bond_yield(self):
        class AkStub:
            def bond_zh_us_rate(self, start_date):
                self.start_date = start_date
                return pd.DataFrame(
                    {
                        "日期": ["2026-07-01", "2026-07-03"],
                        "中国国债收益率10年": ["1.71", "1.7463"],
                    }
                )

        bond = check_signal.fetch_bond_yield(
            AkStub(),
            datetime(2026, 7, 4),
        )

        self.assertEqual(bond.date, "2026-07-03")
        self.assertEqual(bond.yield_10y, 1.7463)
        self.assertIn("AKShare", bond.source)


class SignalEvaluationTest(unittest.TestCase):
    def test_uses_consistent_signal_labels_and_headline(self):
        self.assertEqual(check_signal.evaluate_percentile(85), "A")
        self.assertEqual(check_signal.evaluate_percentile(72), "B")
        self.assertEqual(check_signal.evaluate_percentile(55), "C")
        self.assertEqual(check_signal.evaluate_percentile(30), "D")
        self.assertEqual(check_signal.evaluate_spread(6.0, 1.5), "A")
        self.assertEqual(check_signal.evaluate_spread(4.5, 1.5), "B")
        self.assertEqual(check_signal.evaluate_spread(3.75, 1.5), "C")
        self.assertEqual(check_signal.evaluate_spread(2.9, 1.5), "D")
        self.assertEqual(
            check_signal.build_headline("B", "C", "A"),
            "加大买入区间",
        )
        self.assertEqual(check_signal.normalize_signal_label("历史分位点待验证"), "")


class CsvHeaderLabelTest(unittest.TestCase):
    def test_writes_xlsx_headers_and_bold_signal_columns(self):
        signal = check_signal.Signal(
            run_date="2026-07-06",
            run_time="15:30:00",
            index_date="2026-07-03",
            bond_date="2026-07-03",
            akshare_dividend_yield_2=4.67,
            ytd_return=5.25,
            ytd_max_drawdown=-8.40,
            xueqiu_quote_date="2026-07-06",
            xueqiu_change_percent="-0.73",
            lixinger_date="2026-07-03",
            lixinger_dividend_yield="4.66",
            lixinger_percentile_10y="30.44",
            lixinger_percentile_80_value="6.12",
            lixinger_source="理杏仁公开页面",
            bond_10y_yield=1.7463,
            spread=2.9237,
            percentile_signal="D",
            absolute_signal="B",
            spread_signal="C",
            headline_signal="未进入加大买入区间",
            source_note="test",
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            legacy_csv_path = Path(tmpdir) / "daily.csv"
            legacy_csv_path.write_text(
                "run_date__record_date,index_date__index_valuation_date,"
                "bond_date__china_10y_yield_date\n"
                "2026-07-06,2026-07-03,2026-07-03\n",
                encoding="utf-8-sig",
            )
            xlsx_path = Path(tmpdir) / "daily.xlsx"

            check_signal.write_xlsx(signal, xlsx_path, legacy_csv_path)

            workbook = load_workbook(xlsx_path)
            sheet = workbook.active

        self.assertEqual(
            [cell.value for cell in sheet[1]],
            [
                "run_date__record_date",
                "index_date__index_valuation_date",
                "akshare_dividend_yield_2",
                "ytd_return",
                "ytd_max_drawdown",
                "xueqiu_change_percent",
                "bond_date__china_10y_yield_date",
                "bond_10y_yield",
                "spread",
                "percentile_signal",
                "absolute_signal",
                "spread_signal",
                "headline_signal",
                "lixinger_date__percentile_source_date",
                "lixinger_dividend_yield",
                "lixinger_percentile_10y",
                "lixinger_percentile_80_value",
                "lixinger_source",
            ],
        )
        self.assertEqual(
            [cell.value for cell in sheet[2]],
            [
                "记录日期",
                "指数估值日期",
                "指数股息率口径(%)",
                "全年收益率(%)",
                "年内最大回撤(%)",
                "雪球当天涨跌幅(%)",
                "中国10年国债收益率日期",
                "中国10年国债收益率(%)",
                "股债利差:股息率2-10年国债收益率(百分点)",
                "历史分位点触发",
                "绝对股息率触发",
                "股债利差触发",
                "综合结论",
                "理杏仁估值日期",
                "理杏仁市值加权股息率(%)",
                "理杏仁近10年股息率分位(%)",
                "理杏仁近10年80%分位点(%)",
                "理杏仁数据源状态",
            ],
        )
        run_dates = [
            row[0]
            for row in sheet.iter_rows(min_row=3, values_only=True)
        ]
        self.assertEqual(run_dates.count("2026-07-06"), 1)
        signal_columns = {
            "percentile_signal",
            "absolute_signal",
            "spread_signal",
        }
        header_values = [cell.value for cell in sheet[1]]
        for field in signal_columns:
            column_index = header_values.index(field) + 1
            for row_index in range(1, sheet.max_row + 1):
                self.assertTrue(sheet.cell(row=row_index, column=column_index).font.bold)
        for cell in sheet[sheet.max_row]:
            self.assertEqual(cell.fill.fill_type, None)
        self.assertFalse(sheet.sheet_view.showGridLines)
        self.assertEqual(sheet.freeze_panes, "A3")
        self.assertTrue(sheet.row_dimensions[2].hidden)
        self.assertEqual(sheet.row_dimensions[2].outlineLevel, 1)
        self.assertFalse(sheet.sheet_properties.outlinePr.summaryBelow)
        self.assertIsNotNone(sheet.auto_filter.ref)
        self.assertEqual(sheet["A1"].fill.fgColor.rgb, "001F4E78")
        self.assertEqual(sheet["A2"].fill.fgColor.rgb, "00D9EAF7")
        self.assertGreater(sheet.column_dimensions["A"].width, 9)


if __name__ == "__main__":
    unittest.main()
