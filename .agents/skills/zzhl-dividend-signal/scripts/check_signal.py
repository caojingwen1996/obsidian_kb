#!/usr/bin/env python
"""Check CSI Dividend Index dividend-yield signal and write local records."""

from __future__ import annotations

import argparse
import csv
import gzip
import html
import json
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable
from urllib import request


INDEX_CODE = "000922"
LIXINGER_INDEX_PATH = "sh/000922/922"
LIXINGER_PUBLIC_URL = (
    "https://www.lixinger.com/equity/index/detail/"
    f"{LIXINGER_INDEX_PATH}/fundamental/valuation/dyr?metrics-type=mcw"
)
XUEQIU_SYMBOL = "SH000922"
XUEQIU_REALTIME_URL = (
    "https://stock.xueqiu.com/v5/stock/realtime/quotec.json"
    f"?symbol={XUEQIU_SYMBOL}"
)


@dataclass
class LixingerPercentile:
    date: str
    dividend_yield: float
    percentile_10y: float | None
    percentile_80_value: float | None
    source: str
    note: str


@dataclass
class XueqiuQuote:
    date: str
    change_percent: float | None
    source: str
    note: str


@dataclass
class Signal:
    run_date: str
    run_time: str
    index_date: str
    bond_date: str
    akshare_dividend_yield_2: float
    xueqiu_quote_date: str
    xueqiu_change_percent: str
    lixinger_date: str
    lixinger_dividend_yield: str
    lixinger_percentile_10y: str
    lixinger_percentile_80_value: str
    lixinger_source: str
    bond_10y_yield: float
    spread: float
    percentile_signal: str
    absolute_signal: str
    spread_signal: str
    headline_signal: str
    source_note: str


CSV_HEADER_LABELS = {
    "run_date": "run_date__record_date",
    "run_time": "run_time__record_time",
    "index_date": "index_date__index_valuation_date",
    "bond_date": "bond_date__china_10y_yield_date",
    "lixinger_date": "lixinger_date__percentile_source_date",
}

CSV_CHINESE_LABELS = {
    "run_date": "记录日期",
    "index_date": "指数估值日期",
    "akshare_dividend_yield_2": "AKShare股息率2(%)",
    "xueqiu_change_percent": "雪球当天涨跌幅(%)",
    "bond_date": "中国10年国债收益率日期",
    "bond_10y_yield": "中国10年国债收益率(%)",
    "spread": "股债利差:股息率2-10年国债收益率(百分点)",
    "percentile_signal": "历史分位点触发",
    "absolute_signal": "绝对股息率触发",
    "spread_signal": "股债利差触发",
    "headline_signal": "综合结论",
    "lixinger_date": "理杏仁估值日期",
    "lixinger_dividend_yield": "理杏仁市值加权股息率(%)",
    "lixinger_percentile_10y": "理杏仁近10年股息率分位(%)",
    "lixinger_percentile_80_value": "理杏仁近10年80%分位点(%)",
    "lixinger_source": "理杏仁数据源状态",
}

CSV_FIELD_ORDER = [
    "run_date",
    "index_date",
    "akshare_dividend_yield_2",
    "xueqiu_change_percent",
    "bond_date",
    "bond_10y_yield",
    "spread",
    "percentile_signal",
    "absolute_signal",
    "spread_signal",
    "headline_signal",
    "lixinger_date",
    "lixinger_dividend_yield",
    "lixinger_percentile_10y",
    "lixinger_percentile_80_value",
    "lixinger_source",
]

EXCEL_BOLD_FIELDS = {
    "percentile_signal",
    "absolute_signal",
    "spread_signal",
}
EXCEL_HIGHLIGHT_FILL = "FFF2CC"
EXCEL_HEADER_FILL = "1F4E78"
EXCEL_LABEL_FILL = "D9EAF7"
EXCEL_BORDER_COLOR = "D9E2F3"

SIGNAL_LEVELS = {
    "D": 0,
    "C": 1,
    "B": 2,
    "A": 3,
}

SIGNAL_GRADE_LABELS = {
    "D": "不买或少买",
    "C": "小额定投",
    "B": "加大买入",
    "A": "重点买入",
}

LEGACY_SIGNAL_LABELS = {
    "历史分位点待验证": "",
    "不买或少买": "D",
    "观望": "D",
    "红利吸引力一般": "D",
    "小额定投": "C",
    "中性等待": "C",
    "有配置价值": "C",
    "加大买入": "B",
    "分批买入": "B",
    "红利性价比较高": "B",
    "重点买入": "A",
    "强买入": "A",
    "历史上偏强买点": "A",
}


def import_akshare():
    try:
        import akshare as ak  # type: ignore
    except ModuleNotFoundError as exc:
        raise SystemExit(
            "缺少 AKShare。请在自动化使用的 Python 环境中安装：pip install akshare"
        ) from exc
    return ak


def to_float(value) -> float:
    if value is None:
        raise ValueError("empty numeric value")
    if isinstance(value, str):
        value = value.strip().replace("%", "").replace(",", "")
    return float(value)


def first_existing(columns: Iterable[str], candidates: Iterable[str]) -> str:
    column_set = set(columns)
    for candidate in candidates:
        if candidate in column_set:
            return candidate
    raise KeyError(f"missing expected columns; tried {', '.join(candidates)}")


def select_latest_on_or_before(frame, date_col: str, target_date: datetime):
    dated = frame.copy()
    dated[date_col] = dated[date_col].astype(str)
    dated["_date"] = dated[date_col].apply(lambda value: datetime.fromisoformat(value[:10]))
    dated = dated[dated["_date"] <= target_date].sort_values("_date")
    if dated.empty:
        raise RuntimeError(f"No rows on or before target date {target_date:%Y-%m-%d}")
    return dated.iloc[-1]


def normalize_percentile(value: float | None) -> float | None:
    if value is None:
        return None
    return value * 100 if 0 <= value <= 1 else value


def extract_percent(pattern: str, text: str) -> float | None:
    match = re.search(pattern, text)
    return to_float(match.group(1)) if match else None


def normalize_html_text(text: str) -> str:
    text = html.unescape(text)
    text = re.sub(r"<script\b[^>]*>.*?</script>", " ", text, flags=re.I | re.S)
    text = re.sub(r"<style\b[^>]*>.*?</style>", " ", text, flags=re.I | re.S)
    text = re.sub(r"<[^>]+>", " ", text)
    return re.sub(r"\s+", " ", text)


def parse_lixinger_public_percentile(text: str, source_url: str) -> LixingerPercentile:
    normalized = normalize_html_text(text)
    date_match = re.search(r"最后更新于[:：]?\s*(\d{4}-\d{2}-\d{2})", normalized)
    if date_match is None:
        date_match = re.search(r"更新于\s*(\d{4}-\d{2}-\d{2})", normalized)
    current = extract_percent(r"当前值[:：]?\s*([0-9.]+)%", normalized)
    percentile = extract_percent(r"当前分位点\s*([0-9.]+)%", normalized)
    p80 = extract_percent(r"80%\s*分位点\s*([0-9.]+)%", normalized)
    if current is None or percentile is None or p80 is None:
        raise RuntimeError("无法从理杏仁公开页面解析股息率分位字段")
    return LixingerPercentile(
        date=date_match.group(1) if date_match else "",
        dividend_yield=current,
        percentile_10y=percentile,
        percentile_80_value=p80,
        source="理杏仁公开页面",
        note=source_url,
    )


def fetch_lixinger_public_percentile() -> LixingerPercentile:
    req = request.Request(
        LIXINGER_PUBLIC_URL,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/126.0 Safari/537.36"
            ),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Encoding": "gzip",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            "Referer": "https://www.lixinger.com/",
        },
    )
    with request.urlopen(req, timeout=60) as response:
        raw = response.read()
        if response.headers.get("Content-Encoding") == "gzip":
            raw = gzip.decompress(raw)
        text = raw.decode("utf-8", errors="replace")
    return parse_lixinger_public_percentile(text, LIXINGER_PUBLIC_URL)


def quote_date_from_timestamp(value) -> str:
    if value in (None, ""):
        return ""
    timestamp = float(value)
    if timestamp > 10_000_000_000:
        timestamp = timestamp / 1000
    return datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d")


def parse_xueqiu_quote_payload(text: str) -> XueqiuQuote:
    if not text.strip():
        raise RuntimeError("雪球实时行情接口返回空响应")
    payload = json.loads(text)
    data = payload.get("data")
    if isinstance(data, list):
        quote = data[0] if data else {}
    elif isinstance(data, dict):
        quote = data.get("quote") if isinstance(data.get("quote"), dict) else data
    else:
        quote = {}
    if not isinstance(quote, dict):
        raise RuntimeError("雪球实时行情接口返回结构异常")

    change = None
    for key in ("percent", "chg_percent", "change_percent"):
        if quote.get(key) is not None:
            change = to_float(quote[key])
            break
    if change is None:
        raise RuntimeError("雪球实时行情接口缺少涨跌幅字段")

    quote_date = (
        str(quote.get("date", ""))[:10]
        or quote_date_from_timestamp(quote.get("timestamp"))
        or quote_date_from_timestamp(quote.get("time"))
    )
    return XueqiuQuote(
        date=quote_date,
        change_percent=change,
        source="雪球实时行情",
        note=XUEQIU_REALTIME_URL,
    )


def fetch_xueqiu_realtime_quote() -> XueqiuQuote:
    req = request.Request(
        XUEQIU_REALTIME_URL,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/126.0 Safari/537.36"
            ),
            "Accept": "application/json,text/plain,*/*",
            "Referer": f"https://xueqiu.com/S/{XUEQIU_SYMBOL}",
        },
    )
    with request.urlopen(req, timeout=25) as response:
        text = response.read().decode("utf-8", errors="replace")
    return parse_xueqiu_quote_payload(text)


def get_xueqiu_realtime_quote() -> XueqiuQuote:
    try:
        return fetch_xueqiu_realtime_quote()
    except Exception as exc:
        return XueqiuQuote(
            date="",
            change_percent=None,
            source="雪球实时行情待验证",
            note=f"雪球实时行情失败：{exc}",
        )


def make_llm_lixinger_percentile(
    date: str,
    dividend_yield: str | None,
    percentile_10y: str | None,
    percentile_80_value: str | None,
    source_note: str | None,
) -> LixingerPercentile | None:
    if not (dividend_yield and percentile_10y and percentile_80_value):
        return None
    return LixingerPercentile(
        date=date,
        dividend_yield=to_float(dividend_yield),
        percentile_10y=normalize_percentile(to_float(percentile_10y)),
        percentile_80_value=to_float(percentile_80_value),
        source="大模型网页查询",
        note=source_note or "由 Codex/大模型通过网页查询后传入脚本",
    )


def fetch_lixinger_percentile(
    llm_percentile: LixingerPercentile | None = None,
) -> LixingerPercentile:
    errors: list[str] = []
    try:
        return fetch_lixinger_public_percentile()
    except Exception as public_exc:
        public_error = str(public_exc)
        errors.append(f"公开页面失败：{public_error}")

    if llm_percentile is not None:
        llm_percentile.note = f"{llm_percentile.note}; 公开页面失败后采用：{public_error}"
        return llm_percentile

    errors.append("未提供大模型网页查询结果")

    return LixingerPercentile(
        date="",
        dividend_yield=0.0,
        percentile_10y=None,
        percentile_80_value=None,
        source="理杏仁待验证",
        note="; ".join(errors),
    )


def evaluate_percentile(percentile: float | None) -> str:
    if percentile is None:
        return ""
    if percentile >= 80:
        return "A"
    if percentile >= 70:
        return "B"
    if percentile < 50:
        return "D"
    return "C"


def evaluate_absolute(dividend_yield: float) -> str:
    if dividend_yield < 4.0:
        return "D"
    if dividend_yield < 5.0:
        return "C"
    if dividend_yield <= 6.0:
        return "B"
    return "A"


def evaluate_spread(spread: float) -> str:
    if spread > 3.0:
        return "A"
    if spread > 2.5:
        return "B"
    if spread >= 1.5:
        return "C"
    return "D"


def signal_level(signal: object) -> int:
    return SIGNAL_LEVELS.get(normalize_signal_label(signal), -1)


def normalize_signal_label(signal: object) -> str:
    text = "" if signal is None else str(signal)
    return LEGACY_SIGNAL_LABELS.get(text, text)


def strong_signal_count(signals: Iterable[object]) -> int:
    return sum(1 for signal in signals if signal_level(signal) >= SIGNAL_LEVELS["B"])


def should_highlight_signal_row(row: dict[str, object]) -> bool:
    return (
        strong_signal_count(
            [
                row.get("percentile_signal", ""),
                row.get("absolute_signal", ""),
                row.get("spread_signal", ""),
            ]
        )
        >= 2
    )


def build_headline(
    percentile_signal: str,
    absolute_signal: str,
    spread_signal: str,
) -> str:
    count = strong_signal_count([percentile_signal, absolute_signal, spread_signal])
    if count >= 2:
        return "加大买入区间"
    if not percentile_signal:
        return "历史分位点待验证，暂不判定加大买入区间"
    return "未进入加大买入区间"


def normalize_output_row(row: dict[str, object]) -> dict[str, object]:
    normalized = dict(row)
    for field in ("percentile_signal", "absolute_signal", "spread_signal"):
        normalized[field] = normalize_signal_label(normalized.get(field, ""))
    normalized["headline_signal"] = build_headline(
        str(normalized.get("percentile_signal", "")),
        str(normalized.get("absolute_signal", "")),
        str(normalized.get("spread_signal", "")),
    )
    return normalized


def display_signal_grade(signal: str) -> str:
    if not signal:
        return "待验证"
    return f"{signal}（{SIGNAL_GRADE_LABELS.get(signal, signal)}）"


def fetch_signal(
    llm_percentile: LixingerPercentile | None = None,
    target_run_date: datetime | None = None,
) -> Signal:
    ak = import_akshare()
    valuation = ak.stock_zh_index_value_csindex(symbol=INDEX_CODE)
    if valuation.empty:
        raise RuntimeError(f"AKShare did not return valuation data for {INDEX_CODE}")

    date_col = first_existing(valuation.columns, ["日期", "date"])
    dividend_col = first_existing(
        valuation.columns, ["股息率2", "股息率1", "股息率", "dividend_yield"]
    )

    valuation = valuation.copy()
    valuation["_dividend_yield"] = valuation[dividend_col].apply(to_float)
    run_date = target_run_date or datetime.now()
    latest = select_latest_on_or_before(valuation, date_col, run_date)
    latest_date = latest["_date"]
    index_date = latest_date.strftime("%Y-%m-%d")
    dividend_yield_2 = float(latest["_dividend_yield"])
    xueqiu_quote = get_xueqiu_realtime_quote()

    bond = ak.bond_zh_us_rate(
        start_date=latest_date.replace(year=latest_date.year - 10).strftime("%Y%m%d")
    )
    if bond.empty:
        raise RuntimeError("AKShare did not return China government bond yield data")
    bond_date_col = first_existing(bond.columns, ["日期", "date"])
    bond_rate_col = first_existing(bond.columns, ["中国国债收益率10年", "中国10年期国债"])
    bond = bond.copy()
    bond["_bond_10y_yield"] = bond[bond_rate_col].apply(to_float)
    latest_bond = select_latest_on_or_before(bond, bond_date_col, latest_date)
    bond_10y = float(latest_bond["_bond_10y_yield"])
    spread = dividend_yield_2 - bond_10y

    lixinger = fetch_lixinger_percentile(llm_percentile)
    percentile_signal = evaluate_percentile(lixinger.percentile_10y)
    absolute_signal = evaluate_absolute(dividend_yield_2)
    spread_signal = evaluate_spread(spread)
    headline = build_headline(percentile_signal, absolute_signal, spread_signal)

    return Signal(
        run_date=run_date.strftime("%Y-%m-%d"),
        run_time=datetime.now().strftime("%H:%M:%S"),
        index_date=index_date,
        bond_date=latest_bond["_date"].strftime("%Y-%m-%d"),
        akshare_dividend_yield_2=dividend_yield_2,
        xueqiu_quote_date=xueqiu_quote.date,
        xueqiu_change_percent=(
            ""
            if xueqiu_quote.change_percent is None
            else f"{xueqiu_quote.change_percent:.2f}"
        ),
        lixinger_date=lixinger.date,
        lixinger_dividend_yield=(
            "" if lixinger.dividend_yield <= 0 else f"{lixinger.dividend_yield:.2f}"
        ),
        lixinger_percentile_10y=(
            "" if lixinger.percentile_10y is None else f"{lixinger.percentile_10y:.2f}"
        ),
        lixinger_percentile_80_value=(
            ""
            if lixinger.percentile_80_value is None
            else f"{lixinger.percentile_80_value:.2f}"
        ),
        lixinger_source=lixinger.source,
        bond_10y_yield=bond_10y,
        spread=spread,
        percentile_signal=percentile_signal,
        absolute_signal=absolute_signal,
        spread_signal=spread_signal,
        headline_signal=headline,
        source_note=(
            "AKShare: stock_zh_index_value_csindex(000922, 股息率2), "
            f"bond_zh_us_rate; {xueqiu_quote.source}: {xueqiu_quote.note}; "
            f"{lixinger.source}: {lixinger.note}"
        ),
    )


def output_field_name(name: str) -> str:
    return CSV_HEADER_LABELS.get(name, name)


def signal_output_row(signal: Signal) -> dict[str, object]:
    return normalize_output_row(
        {output_field_name(name): getattr(signal, name) for name in CSV_FIELD_ORDER}
    )


def read_rows_from_csv(csv_path: Path) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    if not csv_path.exists():
        return rows
    with csv_path.open("r", newline="", encoding="utf-8-sig") as file:
        for row in csv.DictReader(file):
            row_run_date = row.get(CSV_HEADER_LABELS["run_date"], "")
            if row_run_date == CSV_CHINESE_LABELS["run_date"]:
                continue
            rows.append(
                normalize_output_row(
                    {
                        output_field_name(name): row.get(output_field_name(name), "")
                        for name in CSV_FIELD_ORDER
                    }
                )
            )
    return rows


def read_rows_from_xlsx(xlsx_path: Path) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    if not xlsx_path.exists():
        return rows
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise SystemExit("缺少 openpyxl。请安装：pip install openpyxl") from exc

    workbook = load_workbook(xlsx_path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    for values in sheet.iter_rows(min_row=3, values_only=True):
        row = dict(zip(headers, values))
        row_run_date = row.get(CSV_HEADER_LABELS["run_date"], "")
        if row_run_date != CSV_CHINESE_LABELS["run_date"]:
            rows.append(
                normalize_output_row(
                    {
                        output_field_name(name): row.get(output_field_name(name), "")
                        for name in CSV_FIELD_ORDER
                    }
                )
            )
    return rows


def write_xlsx(
    signal: Signal,
    xlsx_path: Path,
    legacy_csv_path: Path | None = None,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        raise SystemExit("缺少 openpyxl。请安装：pip install openpyxl") from exc

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = CSV_FIELD_ORDER
    output_fieldnames = [output_field_name(name) for name in fieldnames]
    chinese_labels = [CSV_CHINESE_LABELS.get(name, "") for name in fieldnames]

    rows = read_rows_from_xlsx(xlsx_path)
    if not rows and legacy_csv_path is not None:
        rows = read_rows_from_csv(legacy_csv_path)
    rows = [
        row
        for row in rows
        if row.get(CSV_HEADER_LABELS["run_date"], "") != signal.run_date
    ]
    rows.append(signal_output_row(signal))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "中证红利每日信号"
    sheet.sheet_view.showGridLines = False
    sheet.append(output_fieldnames)
    sheet.append(chinese_labels)
    for row in rows:
        sheet.append([row.get(name, "") for name in output_fieldnames])

    header_fill = PatternFill(fill_type="solid", fgColor=EXCEL_HEADER_FILL)
    label_fill = PatternFill(fill_type="solid", fgColor=EXCEL_LABEL_FILL)
    highlight_fill = PatternFill(
        fill_type="solid",
        fgColor=EXCEL_HIGHLIGHT_FILL,
    )
    thin_border = Border(
        left=Side(style="thin", color=EXCEL_BORDER_COLOR),
        right=Side(style="thin", color=EXCEL_BORDER_COLOR),
        top=Side(style="thin", color=EXCEL_BORDER_COLOR),
        bottom=Side(style="thin", color=EXCEL_BORDER_COLOR),
    )
    center_alignment = Alignment(horizontal="center", vertical="center")
    text_alignment = Alignment(horizontal="left", vertical="center")

    for row_index, row in enumerate(rows, start=3):
        if should_highlight_signal_row(row):
            for column_index in range(1, len(output_fieldnames) + 1):
                sheet.cell(row=row_index, column=column_index).fill = highlight_fill

    bold_columns = {
        output_fieldnames.index(output_field_name(name)) + 1
        for name in EXCEL_BOLD_FIELDS
    }
    for column_index in bold_columns:
        for row_index in range(1, sheet.max_row + 1):
            sheet.cell(row=row_index, column=column_index).font = Font(bold=True)
    for row in sheet.iter_rows(
        min_row=1,
        max_row=sheet.max_row,
        min_col=1,
        max_col=len(output_fieldnames),
    ):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_alignment
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for cell in sheet[2]:
        cell.font = Font(bold=True, color="1F4E78")
        cell.fill = label_fill

    numeric_formats = {
        "akshare_dividend_yield_2": "0.00",
        "xueqiu_change_percent": "0.00",
        "bond_10y_yield": "0.0000",
        "spread": "0.0000",
        "lixinger_dividend_yield": "0.00",
        "lixinger_percentile_10y": "0.00",
        "lixinger_percentile_80_value": "0.00",
    }
    for field, number_format in numeric_formats.items():
        column_index = output_fieldnames.index(output_field_name(field)) + 1
        for row_index in range(3, sheet.max_row + 1):
            sheet.cell(row=row_index, column=column_index).number_format = number_format

    for field in ("headline_signal", "lixinger_source"):
        column_index = output_fieldnames.index(output_field_name(field)) + 1
        for row_index in range(3, sheet.max_row + 1):
            sheet.cell(row=row_index, column=column_index).alignment = text_alignment

    for column_index, field in enumerate(fieldnames, start=1):
        values = [
            str(sheet.cell(row=row_index, column=column_index).value or "")
            for row_index in range(1, sheet.max_row + 1)
        ]
        width = min(max(max(len(value) for value in values) + 2, 10), 28)
        if field in {"headline_signal", "lixinger_source"}:
            width = 24
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    sheet.auto_filter.ref = sheet.dimensions
    sheet.freeze_panes = "A3"
    sheet.row_dimensions[2].hidden = True
    sheet.row_dimensions[2].outlineLevel = 1
    sheet.sheet_properties.outlinePr.summaryBelow = False
    workbook.save(xlsx_path)


def write_markdown(signal: Signal, md_path: Path) -> None:
    md_path.parent.mkdir(parents=True, exist_ok=True)
    percentile_text = (
        f"{float(signal.lixinger_percentile_10y):.2f}%"
        if signal.lixinger_percentile_10y
        else "待验证"
    )
    p80_text = (
        f"{float(signal.lixinger_percentile_80_value):.2f}%"
        if signal.lixinger_percentile_80_value
        else "待验证"
    )
    text = f"""# 中证红利最新股息率信号

- 运行时间：{signal.run_date} {signal.run_time}
- AKShare 指数估值日期：{signal.index_date}
- 理杏仁估值日期：{signal.lixinger_date or "待确认"}
- 10年国债收益率日期：{signal.bond_date}
- AKShare 中证红利股息率2：{signal.akshare_dividend_yield_2:.2f}%
- 雪球当天涨跌幅：{signal.xueqiu_change_percent + "%" if signal.xueqiu_change_percent else "待验证"}
- 理杏仁市值加权股息率：{signal.lixinger_dividend_yield + "%" if signal.lixinger_dividend_yield else "待验证"}
- 理杏仁近10年股息率分位：{percentile_text}
- 理杏仁近10年80%分位点：{p80_text}
- 中国10年国债收益率：{signal.bond_10y_yield:.2f}%
- AKShare 股息率2 - 10年国债收益率：{signal.spread:.2f}%

## 判断

- 历史分位点触发：{display_signal_grade(signal.percentile_signal)}
- 绝对股息率触发：{display_signal_grade(signal.absolute_signal)}
- 相对债券收益率触发：{display_signal_grade(signal.spread_signal)}
- 综合结论：{signal.headline_signal}

## 来源

- {signal.source_note}

## 说明

- 本记录只是按既定规则生成的观察信号，不构成投资建议。
- AKShare `股息率2` 用于绝对股息率和股债利差判断。
- 理杏仁市值加权股息率用于近10年历史分位判断。
- 若数据源字段、接口或口径变化，需要重新核对脚本。
"""
    md_path.write_text(text, encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", default="sources/manual/中证红利信号")
    parser.add_argument("--run-date", default=None, help="Override run date in YYYY-MM-DD")
    parser.add_argument("--llm-lixinger-date", default=None)
    parser.add_argument("--llm-lixinger-dividend-yield", default=None)
    parser.add_argument("--llm-lixinger-percentile-10y", default=None)
    parser.add_argument("--llm-lixinger-percentile-80-value", default=None)
    parser.add_argument("--llm-source-note", default=None)
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    target_run_date = (
        datetime.fromisoformat(args.run_date) if args.run_date else None
    )
    llm_percentile = make_llm_lixinger_percentile(
        date=args.llm_lixinger_date or args.run_date or "",
        dividend_yield=args.llm_lixinger_dividend_yield,
        percentile_10y=args.llm_lixinger_percentile_10y,
        percentile_80_value=args.llm_lixinger_percentile_80_value,
        source_note=args.llm_source_note,
    )
    signal = fetch_signal(
        llm_percentile=llm_percentile,
        target_run_date=target_run_date,
    )
    write_xlsx(
        signal,
        output_dir / "中证红利每日信号.xlsx",
        legacy_csv_path=output_dir / "中证红利每日信号.csv",
    )
    write_markdown(signal, output_dir / "最新信号.md")
    percentile_print = (
        f"{signal.lixinger_percentile_10y}%"
        if signal.lixinger_percentile_10y
        else "待验证"
    )
    print(
        f"{signal.index_date} {signal.headline_signal}："
        f"AKShare股息率2 {signal.akshare_dividend_yield_2:.2f}%，"
        f"理杏仁分位 {percentile_print}，利差 {signal.spread:.2f}%"
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"中证红利信号检查失败：{exc}", file=sys.stderr)
        raise
