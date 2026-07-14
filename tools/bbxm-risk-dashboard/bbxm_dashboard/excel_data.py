from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from .errors import DashboardDataError

REQUIRED_COLUMNS = ("日期", "当日累计风险提示次数", "风险原因")


def _parse_date(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        return datetime.strptime(value.strip(), "%Y-%m-%d").date().isoformat()
    raise ValueError("日期必须是 Excel 日期或 YYYY-MM-DD 文本")


def load_risk_records(path: Path) -> tuple[list[dict], list[str]]:
    if not path.exists():
        raise DashboardDataError("excel_missing", f"未找到风险记录文件：{path}")
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except PermissionError as exc:
        raise DashboardDataError(
            "excel_locked",
            "Excel 文件正被占用，请关闭后重试",
        ) from exc

    sheet = workbook.active
    header_values = [
        cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))
    ]
    missing = [name for name in REQUIRED_COLUMNS if name not in header_values]
    if missing:
        workbook.close()
        detected = "、".join(str(value) for value in header_values if value)
        raise DashboardDataError(
            "excel_missing_columns",
            f"Excel 缺少必需列：{'、'.join(missing)}；检测到：{detected}",
        )

    positions = {name: header_values.index(name) for name in REQUIRED_COLUMNS}
    grouped: OrderedDict[str, dict] = OrderedDict()
    issues = []
    for row_number, cells in enumerate(
        sheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        raw_date = cells[positions["日期"]]
        raw_count = cells[positions["当日累计风险提示次数"]]
        raw_reason = cells[positions["风险原因"]]
        try:
            parsed_date = _parse_date(raw_date)
        except (TypeError, ValueError):
            issues.append({"row": row_number, "column": "日期", "value": raw_date})
            continue
        if (
            isinstance(raw_count, bool)
            or not isinstance(raw_count, int)
            or raw_count < 0
        ):
            issues.append(
                {
                    "row": row_number,
                    "column": "当日累计风险提示次数",
                    "value": raw_count,
                }
            )
            continue

        reason = "" if raw_reason is None else str(raw_reason).strip()
        item = grouped.setdefault(
            parsed_date,
            {"count": 0, "reasons": [], "rows": 0},
        )
        item["count"] += raw_count
        item["rows"] += 1
        if reason:
            item["reasons"].append(reason)

    workbook.close()
    if issues:
        raise DashboardDataError("excel_invalid", "Excel 中存在无效数据", issues)

    warnings = [
        f"{key} 存在 {item['rows']} 行记录，已合并"
        for key, item in grouped.items()
        if item["rows"] > 1
    ]
    records = [
        {
            "date": key,
            "count": item["count"],
            "reason": "\n".join(item["reasons"]),
        }
        for key, item in sorted(grouped.items())
    ]
    if not records:
        raise DashboardDataError("excel_empty", "Excel 中没有有效风险记录")
    return records, warnings
