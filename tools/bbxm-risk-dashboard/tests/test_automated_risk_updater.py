import importlib.util
import json
from copy import deepcopy
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook


SCRIPT_PATH = (
    Path(__file__).resolve().parents[1]
    / "scripts"
    / "upsert_automated_risk.py"
)
SPEC = importlib.util.spec_from_file_location("upsert_automated_risk", SCRIPT_PATH)
updater = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(updater)

HEADERS = ["日期", "当日累计风险提示次数", "风险原因"]


def make_book(tmp_path, rows):
    path = tmp_path / "risk.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    return path


def read_rows(path):
    workbook = load_workbook(path, data_only=True)
    rows = [list(row) for row in workbook.active.iter_rows(min_row=2, values_only=True)]
    workbook.close()
    return rows


def risk(post_key="post-1", level="R2", reason="承接风险上升"):
    return {
        "post_key": post_key,
        "source_file": f"{post_key}.md",
        "url": f"https://xueqiu.com/{post_key}",
        "title": f"标题-{post_key}",
        "published_at": "2026-07-14T18:00:00+08:00",
        "level": level,
        "risk_object": "市场情绪",
        "trigger": "一致预期增强",
        "transmission": "一致预期增强 → 后手集中 → 承接脆弱",
        "evidence": ["帖子证据"],
        "reason": reason,
    }


def valid_analysis(qualified=None, not_written=None):
    qualified = [] if qualified is None else qualified
    not_written = [] if not_written is None else not_written
    analyzed = len({item["post_key"] for item in qualified + not_written})
    return {
        "schema_version": 1,
        "target_date": "2026-07-14",
        "author": "冰冰小美",
        "generated_at": "2026-07-14T18:30:00+08:00",
        "analysis_complete": True,
        "coverage": {
            "saved_post_count": analyzed,
            "analyzed_post_count": analyzed,
            "unresolved_post_count": 0,
            "unresolved_reasons": [],
        },
        "qualified": qualified,
        "not_written": not_written,
    }


def write_json(path, data):
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return path


def run_update(tmp_path, data, rows, replace_file=None):
    workbook_path = make_book(tmp_path, rows)
    analysis_path = write_json(tmp_path / "risk-analysis.json", data)
    status_path = tmp_path / "risk-write-status.json"
    kwargs = {} if replace_file is None else {"replace_file": replace_file}
    result = updater.update_workbook(
        analysis_path,
        workbook_path,
        status_path,
        **kwargs,
    )
    status = json.loads(status_path.read_text(encoding="utf-8"))
    return result, status, workbook_path


def test_upsert_writes_one_automatic_row_and_preserves_manual_row(tmp_path):
    data = valid_analysis([risk("post-1", "R2"), risk("post-2", "R1")])

    result, status, workbook_path = run_update(
        tmp_path,
        data,
        [["2026-07-14", 1, "人工记录"]],
    )

    rows = read_rows(workbook_path)
    assert result["status"] == "written"
    assert result["count"] == 2
    assert status == result
    assert rows[0] == ["2026-07-14", 1, "人工记录"]
    assert rows[1][0:2] == ["2026-07-14", 2]
    assert rows[1][2].startswith(updater.AUTOMATED_PREFIX)
    assert "【R2】标题-post-1：承接风险上升" in rows[1][2]
    assert "【R1】标题-post-2：承接风险上升" in rows[1][2]


def test_same_day_rerun_replaces_automatic_row_instead_of_appending(tmp_path):
    old_reason = f"{updater.AUTOMATED_PREFIX}\n旧风险"
    data = valid_analysis([risk("post-new", "R3", "风险共振")])

    result, _, workbook_path = run_update(
        tmp_path,
        data,
        [
            ["2026-07-14", 1, "人工记录"],
            ["2026-07-14", 3, old_reason],
            ["2026-07-13", 1, f"{updater.AUTOMATED_PREFIX}\n其他日期"],
        ],
    )

    rows = read_rows(workbook_path)
    assert result["status"] == "written"
    assert len(rows) == 3
    assert rows[0] == ["2026-07-14", 1, "人工记录"]
    assert rows[1] == ["2026-07-13", 1, f"{updater.AUTOMATED_PREFIX}\n其他日期"]
    assert rows[2][0:2] == ["2026-07-14", 1]
    assert "风险共振" in rows[2][2]


def test_empty_qualified_removes_only_existing_automatic_row(tmp_path):
    data = valid_analysis(
        not_written=[{"post_key": "post-1", "level": "N", "reason": "不构成风险"}]
    )
    old_reason = f"{updater.AUTOMATED_PREFIX}\n旧风险"

    result, _, workbook_path = run_update(
        tmp_path,
        data,
        [
            ["2026-07-14", 1, "人工记录"],
            ["2026-07-14", 2, old_reason],
            ["2026-07-13", 1, old_reason],
        ],
    )

    assert result["status"] == "removed"
    assert result["count"] == 0
    assert read_rows(workbook_path) == [
        ["2026-07-14", 1, "人工记录"],
        ["2026-07-13", 1, old_reason],
    ]


def test_empty_qualified_without_existing_row_returns_no_risk(tmp_path):
    data = valid_analysis(
        not_written=[{"post_key": "post-1", "level": "W1", "reason": "风险转弱"}]
    )

    result, status, workbook_path = run_update(
        tmp_path,
        data,
        [["2026-07-14", 1, "人工记录"]],
    )

    assert result["status"] == "no_risk"
    assert status["count"] == 0
    assert read_rows(workbook_path) == [["2026-07-14", 1, "人工记录"]]


def test_duplicate_equal_post_keys_count_once(tmp_path):
    item = risk("post-1", "R2")
    data = valid_analysis([item, deepcopy(item)])

    result, _, workbook_path = run_update(tmp_path, data, [])

    assert result["status"] == "written"
    assert result["count"] == 1
    assert read_rows(workbook_path)[0][1] == 1


@pytest.mark.parametrize(
    "mutation",
    [
        lambda data: data.update(analysis_complete=False),
        lambda data: data["coverage"].update(saved_post_count=3),
        lambda data: data["qualified"][0].update(level="N"),
        lambda data: data["qualified"][0].update(source_file="", url=""),
        lambda data: data["qualified"][0].update(transmission=""),
    ],
)
def test_invalid_or_incomplete_analysis_is_blocked_without_excel_change(
    tmp_path,
    mutation,
):
    data = valid_analysis([risk("post-1", "R2")])
    mutation(data)
    workbook_path = make_book(tmp_path, [["2026-07-14", 1, "人工记录"]])
    original = workbook_path.read_bytes()
    analysis_path = write_json(tmp_path / "risk-analysis.json", data)
    status_path = tmp_path / "risk-write-status.json"

    result = updater.update_workbook(analysis_path, workbook_path, status_path)

    assert result["status"] == "blocked"
    assert result["count"] == 0
    assert workbook_path.read_bytes() == original
    assert json.loads(status_path.read_text(encoding="utf-8")) == result


def test_duplicate_conflicting_post_keys_are_blocked(tmp_path):
    first = risk("post-1", "R1")
    second = risk("post-1", "R2")
    data = valid_analysis([first, second])

    result, _, workbook_path = run_update(
        tmp_path,
        data,
        [["2026-07-14", 1, "人工记录"]],
    )

    assert result["status"] == "blocked"
    assert read_rows(workbook_path) == [["2026-07-14", 1, "人工记录"]]


def test_replace_failure_returns_pending_and_preserves_original(tmp_path):
    data = valid_analysis([risk("post-1", "R2")])

    def fail_replace(_source, _destination):
        raise PermissionError("workbook locked")

    workbook_path = make_book(tmp_path, [["2026-07-14", 1, "人工记录"]])
    original = workbook_path.read_bytes()
    analysis_path = write_json(tmp_path / "risk-analysis.json", data)
    status_path = tmp_path / "risk-write-status.json"

    result = updater.update_workbook(
        analysis_path,
        workbook_path,
        status_path,
        replace_file=fail_replace,
    )

    assert result["status"] == "pending"
    assert result["count"] == 1
    assert workbook_path.read_bytes() == original
    assert json.loads(status_path.read_text(encoding="utf-8")) == result


def test_status_contains_traceable_paths_and_target_date(tmp_path):
    data = valid_analysis([risk("post-1", "R1")])

    result, _, _ = run_update(tmp_path, data, [])

    assert result["target_date"] == "2026-07-14"
    assert result["analysis_file"].endswith("risk-analysis.json")
    assert result["workbook"].endswith("risk.xlsx")
    assert result["message"]
