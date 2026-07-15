from __future__ import annotations

import argparse
import json
import os
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook


AUTOMATED_PREFIX = "[自动分析｜冰冰小美每日任务]"
REQUIRED_COLUMNS = ("日期", "当日累计风险提示次数", "风险原因")
ALLOWED_LEVELS = {"R1", "R2", "R3", "W1", "W2", "W3"}
NOT_WRITTEN_LEVELS = {"N", "待验证"}


class AnalysisError(ValueError):
    pass


def _required_text(item: dict, name: str) -> str:
    value = item.get(name)
    if not isinstance(value, str) or not value.strip():
        raise AnalysisError(f"{name} 必须是非空文本")
    return value.strip()


def _parse_iso_date(value: object, name: str) -> str:
    if not isinstance(value, str):
        raise AnalysisError(f"{name} 必须是 YYYY-MM-DD 文本")
    try:
        parsed = date.fromisoformat(value)
    except ValueError as exc:
        raise AnalysisError(f"{name} 必须是有效的 YYYY-MM-DD 日期") from exc
    if parsed.isoformat() != value:
        raise AnalysisError(f"{name} 必须使用 YYYY-MM-DD 格式")
    return value


def _parse_iso_datetime(value: object, name: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise AnalysisError(f"{name} 必须是非空 ISO 时间")
    try:
        parsed = datetime.fromisoformat(value)
    except ValueError as exc:
        raise AnalysisError(f"{name} 必须是有效的 ISO 时间") from exc
    if "T" not in value or parsed.utcoffset() is None:
        raise AnalysisError(f"{name} 必须包含时间和 UTC 偏移")
    return parsed.isoformat()


def _non_negative_int(value: object, name: str) -> int:
    if isinstance(value, bool) or not isinstance(value, int) or value < 0:
        raise AnalysisError(f"{name} 必须是大于或等于 0 的整数")
    return value


def _canonical(item: dict) -> str:
    return json.dumps(item, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def load_and_validate_analysis(path: Path) -> dict:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise AnalysisError(f"无法读取风险分析文件：{exc}") from exc

    if not isinstance(data, dict):
        raise AnalysisError("风险分析根节点必须是对象")
    if data.get("schema_version") != 1:
        raise AnalysisError("schema_version 必须为 1")

    data["target_date"] = _parse_iso_date(data.get("target_date"), "target_date")
    if data.get("author") != "冰冰小美":
        raise AnalysisError("author 必须为冰冰小美")
    _parse_iso_datetime(data.get("generated_at"), "generated_at")
    if data.get("analysis_complete") is not True:
        raise AnalysisError("analysis_complete=false，禁止更新 Excel")

    coverage = data.get("coverage")
    if not isinstance(coverage, dict):
        raise AnalysisError("coverage 必须是对象")
    saved = _non_negative_int(coverage.get("saved_post_count"), "saved_post_count")
    analyzed = _non_negative_int(
        coverage.get("analyzed_post_count"),
        "analyzed_post_count",
    )
    unresolved = _non_negative_int(
        coverage.get("unresolved_post_count"),
        "unresolved_post_count",
    )
    unresolved_reasons = coverage.get("unresolved_reasons")
    if not isinstance(unresolved_reasons, list) or not all(
        isinstance(item, str) for item in unresolved_reasons
    ):
        raise AnalysisError("unresolved_reasons 必须是文本数组")
    if saved != analyzed + unresolved:
        raise AnalysisError("saved_post_count 必须等于 analyzed_post_count + unresolved_post_count")
    if unresolved != 0:
        raise AnalysisError("analysis_complete=true 时 unresolved_post_count 必须为 0")

    qualified = data.get("qualified")
    not_written = data.get("not_written")
    if not isinstance(qualified, list) or not isinstance(not_written, list):
        raise AnalysisError("qualified 和 not_written 必须是数组")

    seen: dict[str, tuple[str, str]] = {}
    unique_qualified = []
    for item in qualified:
        if not isinstance(item, dict):
            raise AnalysisError("qualified 条目必须是对象")
        post_key = _required_text(item, "post_key")
        level = _required_text(item, "level")
        if level not in ALLOWED_LEVELS:
            raise AnalysisError(f"qualified 等级无效：{level}")
        for name in ("title", "published_at", "risk_object", "trigger", "transmission", "reason"):
            _required_text(item, name)
        _parse_iso_datetime(item["published_at"], "published_at")
        source_file = item.get("source_file")
        url = item.get("url")
        if not any(isinstance(value, str) and value.strip() for value in (source_file, url)):
            raise AnalysisError("qualified 条目必须提供 source_file 或 url")
        evidence = item.get("evidence")
        if not isinstance(evidence, list) or not evidence or not all(
            isinstance(value, str) and value.strip() for value in evidence
        ):
            raise AnalysisError("evidence 必须是非空文本数组")

        fingerprint = _canonical(item)
        previous = seen.get(post_key)
        if previous and previous != ("qualified", fingerprint):
            raise AnalysisError(f"post_key 存在矛盾内容：{post_key}")
        if not previous:
            seen[post_key] = ("qualified", fingerprint)
            unique_qualified.append(item)

    for item in not_written:
        if not isinstance(item, dict):
            raise AnalysisError("not_written 条目必须是对象")
        post_key = _required_text(item, "post_key")
        level = _required_text(item, "level")
        if level not in NOT_WRITTEN_LEVELS:
            raise AnalysisError(f"not_written 等级无效：{level}")
        _required_text(item, "reason")
        fingerprint = _canonical(item)
        previous = seen.get(post_key)
        if previous and previous != ("not_written", fingerprint):
            raise AnalysisError(f"post_key 存在矛盾内容：{post_key}")
        if not previous:
            seen[post_key] = ("not_written", fingerprint)

    if analyzed != len(seen):
        raise AnalysisError("analyzed_post_count 必须等于唯一 post_key 数量")

    data["qualified"] = unique_qualified
    return data


def format_automatic_reason(items: list[dict]) -> str:
    lines = [AUTOMATED_PREFIX]
    for item in items:
        source = item.get("url") or item.get("source_file")
        lines.append(
            f"【{item['level']}】{item['title']}：{item['reason']}（来源：{source}）"
        )
    return "\n".join(lines)


def _status_payload(
    analysis_path: Path,
    workbook_path: Path,
    target_date: str,
    status: str,
    count: int,
    message: str,
) -> dict:
    return {
        "schema_version": 1,
        "target_date": target_date,
        "status": status,
        "count": count,
        "workbook": str(workbook_path),
        "analysis_file": str(analysis_path),
        "message": message,
    }


def _write_status(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_suffix(path.suffix + ".tmp")
    temp_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    os.replace(temp_path, path)


def _cell_date(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        try:
            return date.fromisoformat(value.strip()).isoformat()
        except ValueError:
            return None
    return None


def _save_and_replace(
    workbook,
    workbook_path: Path,
    replace_file: Callable[[str | Path, str | Path], object],
) -> None:
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    handle = tempfile.NamedTemporaryFile(
        prefix=f".{workbook_path.stem}-",
        suffix=workbook_path.suffix,
        dir=workbook_path.parent,
        delete=False,
    )
    temp_path = Path(handle.name)
    handle.close()
    try:
        workbook.save(temp_path)
        replace_file(temp_path, workbook_path)
    finally:
        temp_path.unlink(missing_ok=True)


def update_workbook(
    analysis_path: Path,
    workbook_path: Path,
    status_path: Path,
    *,
    replace_file: Callable[[str | Path, str | Path], object] = os.replace,
) -> dict:
    analysis_path = Path(analysis_path)
    workbook_path = Path(workbook_path)
    status_path = Path(status_path)
    raw_target_date = ""
    try:
        raw = json.loads(analysis_path.read_text(encoding="utf-8"))
        if isinstance(raw, dict) and isinstance(raw.get("target_date"), str):
            raw_target_date = raw["target_date"]
    except (OSError, UnicodeError, json.JSONDecodeError):
        pass

    try:
        analysis = load_and_validate_analysis(analysis_path)
    except AnalysisError as exc:
        payload = _status_payload(
            analysis_path,
            workbook_path,
            raw_target_date,
            "blocked",
            0,
            str(exc),
        )
        _write_status(status_path, payload)
        return payload

    target_date = analysis["target_date"]
    qualified = analysis["qualified"]
    try:
        workbook = load_workbook(workbook_path)
    except PermissionError as exc:
        payload = _status_payload(
            analysis_path,
            workbook_path,
            target_date,
            "pending",
            len(qualified),
            f"Excel 被占用，等待同日重跑：{exc}",
        )
        _write_status(status_path, payload)
        return payload
    except (OSError, ValueError) as exc:
        payload = _status_payload(
            analysis_path,
            workbook_path,
            target_date,
            "blocked",
            0,
            f"无法读取 Excel：{exc}",
        )
        _write_status(status_path, payload)
        return payload

    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    missing = [name for name in REQUIRED_COLUMNS if name not in headers]
    if missing:
        workbook.close()
        payload = _status_payload(
            analysis_path,
            workbook_path,
            target_date,
            "blocked",
            0,
            f"Excel 缺少必需列：{'、'.join(missing)}",
        )
        _write_status(status_path, payload)
        return payload

    positions = {name: headers.index(name) + 1 for name in REQUIRED_COLUMNS}
    rows_to_delete = []
    for row_number in range(2, sheet.max_row + 1):
        row_date = _cell_date(sheet.cell(row_number, positions["日期"]).value)
        reason = sheet.cell(row_number, positions["风险原因"]).value
        if (
            row_date == target_date
            and isinstance(reason, str)
            and reason.strip().startswith(AUTOMATED_PREFIX)
        ):
            rows_to_delete.append(row_number)

    if not qualified and not rows_to_delete:
        workbook.close()
        payload = _status_payload(
            analysis_path,
            workbook_path,
            target_date,
            "no_risk",
            0,
            "完整分析后没有可写入节点，未修改 Excel",
        )
        _write_status(status_path, payload)
        return payload

    for row_number in reversed(rows_to_delete):
        sheet.delete_rows(row_number)

    if qualified:
        row = [None] * len(headers)
        row[positions["日期"] - 1] = target_date
        row[positions["当日累计风险提示次数"] - 1] = len(qualified)
        row[positions["风险原因"] - 1] = format_automatic_reason(qualified)
        sheet.append(row)

    try:
        _save_and_replace(workbook, workbook_path, replace_file)
    except OSError as exc:
        payload = _status_payload(
            analysis_path,
            workbook_path,
            target_date,
            "pending",
            len(qualified),
            f"Excel 暂时无法替换，等待同日重跑：{exc}",
        )
        _write_status(status_path, payload)
        return payload
    finally:
        workbook.close()

    if qualified:
        status = "written"
        message = "已更新自动分析风险提示行"
    else:
        status = "removed"
        message = "完整分析后没有可写入节点，已删除旧自动分析行"
    payload = _status_payload(
        analysis_path,
        workbook_path,
        target_date,
        status,
        len(qualified),
        message,
    )
    _write_status(status_path, payload)
    return payload


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="幂等更新冰冰小美风险提示 Excel 自动行")
    parser.add_argument("--analysis-file", required=True, type=Path)
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--status-file", required=True, type=Path)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    result = update_workbook(args.analysis_file, args.workbook, args.status_file)
    print(json.dumps(result, ensure_ascii=False))
    return 0 if result["status"] in {"written", "no_risk", "removed"} else 2


if __name__ == "__main__":
    raise SystemExit(main())
